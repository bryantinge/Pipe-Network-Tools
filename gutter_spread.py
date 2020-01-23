import io
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from utils import borders, set_border, set_format


def spread_format(s3, S3_BUCKET, s3_keys_csv, folder_name,
                  spread_names, bypass_toggle):
    '''Formats gutter spread input csv files to an xlsx file'''

    # Create list of series names from input files
    series_names = [re.sub('.txt', '', name) for name in spread_names]
    series_names = [re.sub('_', ' ', name) for name in series_names]

    # Read txt files from s3
    s3_objects = [s3.get_object(Bucket=S3_BUCKET, Key=key)
                  for key in s3_keys_csv]
    spread_files = [obj['Body'] for obj in s3_objects]

    # Create and format dataframes
    dfs = {}
    try:
        for i, file in enumerate(spread_files):
            df = pd.read_csv(
                file,
                sep='\t',
                header=0,
                names=['line', 'structure', 'inlet_type', 'bypass', 'area',
                       'tc', 'intensity', 'c_value', 'flow_inlet',
                       'flow_bypass', 'flow_captured', 'flow_bypassed',
                       'slope', 'spread']
            )

            df.replace({'Outfall': 'OUT', 'Curb': 'CURB', 'Grate': 'GRATE',
                        'Comb.': 'COMB', 'Generic': 'GENERIC', 'Hdwall': 'FES',
                        'None': 'NONE', 'Dp-Curb': 'DP-CURB',
                        'Dp-Grate': 'DP-GRATE', 'Null Structure': 'NONE'},
                       inplace=True)
            df.replace({'Notes:  j-Line contains hyd. jump': ''},
                       inplace=True, regex=True)
            df.replace({' j': '', '\(': '', '\)': '', ' DOUBLE': ''},
                       inplace=True, regex=True)

            df.dropna(axis=0, inplace=True)
            for col in range(4, 14):
                df.iloc[:, col] = pd.to_numeric(
                    df.iloc[:, col], errors='coerce')

            df.loc[df.inlet_type == 'GRATE', 'slope'] = 'N/A'
            df.loc[df.inlet_type == 'GRATE', 'spread'] = 'N/A'
            df.loc[df.bypass == 'SAG', 'slope'] = 'SAG'

            df.drop(df.columns[0], axis=1, inplace=True)
            if not bypass_toggle:
                df.drop(['bypass', 'flow_bypass', 'flow_captured',
                         'flow_bypassed'], axis=1, inplace=True)

            series = series_names[i]
            dfs[series] = df

    except pd.errors.ParserError:
        return None

    except ValueError:
        return None

    # Create unformatted excel file from dataframes
    df_stream = io.BytesIO()
    writer = pd.ExcelWriter(df_stream, engine='openpyxl')
    for series, df in dfs.items():
        df.to_excel(writer, sheet_name=series,
                    startrow=3, startcol=1, index=False)
    writer.save()

    # Read workbook into openpyxl from binary
    wb = load_workbook(io.BytesIO(df_stream.getvalue()))

    # Format excel sheets
    for i, ws in enumerate(wb.worksheets):

        max_row = ws.max_row
        max_col = ws.max_column

        ws['B2'] = series_names[i] + ' SERIES (GUTTER SPREAD)'
        ws['B3'] = 'STRUCTURE'
        ws['C3'] = 'INLET\nTYPE'

        ws['B4'] = ''
        ws['C4'] = ''

        if bypass_toggle:
            ws['D3'] = 'BYPASS\nSTRUCTURE'
            ws['E3'] = 'DRAINAGE\nAREA'
            ws['F3'] = 'TC'
            ws['G3'] = 'I'
            ws['H3'] = 'C'
            ws['I3'] = 'Q\n(INLET)'
            ws['J3'] = 'Q\n(BYPASS)'
            ws['K3'] = 'Q\n(CAPTURED)'
            ws['L3'] = 'Q\n(BYPASSED)'
            ws['M3'] = 'LONG\nSLOPE'
            ws['N3'] = 'GUTTER\nSPREAD'

            ws['D4'] = ''
            ws['E4'] = '(AC)'
            ws['F4'] = '(MIN)'
            ws['G4'] = '(IN/HR)'
            ws['H4'] = ''
            ws['I4'] = '(CFS)'
            ws['J4'] = '(CFS)'
            ws['K4'] = '(CFS)'
            ws['L4'] = '(CFS)'
            ws['M4'] = '(FT/FT)'
            ws['N4'] = '(FT)'

        else:
            ws['D3'] = 'DRAINAGE\nAREA'
            ws['E3'] = 'TC'
            ws['F3'] = 'I'
            ws['G3'] = 'C'
            ws['H3'] = 'Q\n(INLET)'
            ws['I3'] = 'LONG\nSLOPE'
            ws['J3'] = 'GUTTER\nSPREAD'

            ws['D4'] = '(AC)'
            ws['E4'] = '(MIN)'
            ws['F4'] = '(IN/HR)'
            ws['G4'] = ''
            ws['H4'] = '(CFS)'
            ws['I4'] = '(FT/FT)'
            ws['J4'] = '(FT)'

        # Merge cells
        if bypass_toggle:
            ws.merge_cells('B2:N2')
        else:
            ws.merge_cells('B2:J2')

        # Set cell font
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrapText=True)
                cell.font = Font(size=10, name='Arial')

        for row in ws.iter_rows(min_row=2, max_row=4,
                                min_col=2, max_col=max_col):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='Arial')

        # Set cell background fill
        ws['B2'].fill = PatternFill(fgColor='BFBFBF', bgColor='BFBFBF',
                                    fill_type='solid')

        for row in ws.iter_rows(min_row=3, max_row=4,
                                min_col=2, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill(fgColor='D9D9D9', bgColor='D9D9D9',
                                        fill_type='solid')

        # Set cell border
        set_border(ws, borders.thin, 3, max_row, 2, max_col)
        set_border(ws, borders.med_top_left_bot, 2, 2, 2, 2)
        set_border(ws, borders.med_top_right_bot, 2, 2, max_col, max_col)
        set_border(ws, borders.med_tlcorner, 3, 3, 2, 2)
        set_border(ws, borders.med_left, 4, 4, 2, 2)
        set_border(ws, borders.med_trcorner, 3, 3, max_col, max_col)
        set_border(ws, borders.med_right, 4, 4, max_col, max_col)
        set_border(ws, borders.med_tlcorner, 5, 5, 2, 2)
        set_border(ws, borders.med_trcorner, 5, 5, max_col, max_col)
        set_border(ws, borders.med_top_bot, 2, 2, 3, max_col - 1)
        set_border(ws, borders.med_top, 5, 5, 3, max_col - 1)

        # Set cell border for single entry
        if max_row == 5:
            set_border(ws, borders.med_top_left_bot,
                       max_row, max_row, 2, 2)
            set_border(ws, borders.med_top_right_bot,
                       max_row, max_row, max_col, max_col)
            set_border(ws, borders.med_top_bot,
                       max_row, max_row, 3, max_col - 1)

        else:
            set_border(ws, borders.med_left,
                       6, max_row, 2, 2)
            set_border(ws, borders.med_right,
                       6, max_row, max_col, max_col)
            set_border(ws, borders.med_bot,
                       max_row, max_row, 3, max_col - 1)
            set_border(ws, borders.med_blcorner,
                       max_row, max_row, 2, 2)
            set_border(ws, borders.med_brcorner,
                       max_row, max_row, max_col, max_col)

        # Set cell number format
        if bypass_toggle:
            set_format(ws, '0.00', 5, max_row, 5, 5)
            set_format(ws, '0.0', 5, max_row, 6, 6)
            set_format(ws, '0.00', 5, max_row, 7, max_col)

        else:
            set_format(ws, '0.00', 5, max_row, 4, 4)
            set_format(ws, '0.0', 5, max_row, 5, 5)
            set_format(ws, '0.00', 5, max_row, 6, max_col)

        # Set row and column dimensions
        ws.row_dimensions[1].height = 13.8
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 36
        ws.row_dimensions[4].height = 21

        ws.column_dimensions['A'].width = 4.02
        ws.column_dimensions['B'].width = 13.98
        ws.column_dimensions['C'].width = 11.24
        ws.column_dimensions['D'].width = 15.36
        ws.column_dimensions['E'].width = 13.24
        ws.column_dimensions['F'].width = 11.13
        ws.column_dimensions['G'].width = 11.13
        ws.column_dimensions['H'].width = 11.13
        ws.column_dimensions['I'].width = 13.47
        ws.column_dimensions['J'].width = 13.47
        ws.column_dimensions['K'].width = 13.47
        ws.column_dimensions['L'].width = 13.47
        ws.column_dimensions['M'].width = 13.47
        ws.column_dimensions['N'].width = 13.47

    # Convert formatted xlsx file to binary
    stream_xlsx_form = io.BytesIO()
    wb.save(stream_xlsx_form)
    stream_xlsx_form.seek(0)

    # Write formatted xlsx file to s3
    xlsx_form = stream_xlsx_form.getvalue()
    s3_key_xlsx = ''.join(['spread', '/', folder_name, '/',
                           'xlsx', '/', 'Gutter Spread.xlsx'])
    s3.put_object(Bucket=S3_BUCKET, Key=s3_key_xlsx, Body=xlsx_form)

    # Return s3 key
    response = s3_key_xlsx
    return response
