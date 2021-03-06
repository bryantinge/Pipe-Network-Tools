import io
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from utils import borders, set_border, set_format


def design_format(s3, S3_BUCKET, s3_keys_csv, folder_name,
                  design_names, hgl_toggle):
    '''Formats pipe design input csv files to an xlsx file'''

    # Create list of series names from input files
    series_names = [re.sub('.txt', '', name) for name in design_names]
    series_names = [re.sub('_', ' ', name) for name in series_names]

    # Read txt files from s3
    s3_objects = [s3.get_object(Bucket=S3_BUCKET, Key=key)
                  for key in s3_keys_csv]
    design_files = [obj['Body'] for obj in s3_objects]

    # Create and format dataframes
    dfs = {}
    try:
        for i, file in enumerate(design_files):
            df = pd.read_csv(
                file,
                sep='\t',
                header=0,
                names=['line', 'inlet_type', 'struc_from', 'struc_to', 'area',
                       'tc', 'intensity', 'c_value', 'flow_inlet',
                       'flow_total', 'flow_cap', 'length', 'size', 'material',
                       'slope', 'inv_up', 'inv_down', 'rim_up', 'rim_down',
                       'hgl_up', 'hgl_down']
            )

            df.replace({'Outfall': 'OUT', 'Curb': 'CURB', 'Grate': 'GRATE',
                        'Comb.': 'COMB', 'Generic': 'GENERIC', 'Hdwall': 'FES',
                        'None': 'NONE', 'Dp-Curb': 'DP-CURB',
                        'Dp-Grate': 'DP-GRATE', 'Null Structure': 'NONE'},
                       inplace=True)
            df.replace({'Notes:  j-Line contains hyd. jump': ''},
                       inplace=True, regex=True)
            df.replace({' j': '', '\(': '', '\)': '', ' DOUBLE': ''},
                       inplace=True, regex=True)

            df.dropna(axis=0, inplace=True)
            for col in range(4, 21):
                df.iloc[:, col] = pd.to_numeric(
                    df.iloc[:, col], errors='coerce')
            df.drop(df.columns[0], axis=1, inplace=True)
            if not hgl_toggle:
                df.drop(['hgl_up', 'hgl_down'], axis=1, inplace=True)

            series = series_names[i]
            dfs[series] = df

    except pd.errors.ParserError:
        print('ParserError')
        return None

    except ValueError:
        print('ValueError')
        return None

    # Create unformatted excel file from dataframes
    df_stream = io.BytesIO()
    writer = pd.ExcelWriter(df_stream, engine='openpyxl')
    for series, df in dfs.items():
        df.to_excel(writer, sheet_name=series,
                    startrow=3, startcol=1, index=False)
    writer.save()

    # Read workbook into openpyxl from binary
    wb = load_workbook(io.BytesIO(df_stream.getvalue()))

    # Format excel sheets
    for i, ws in enumerate(wb.worksheets):

        max_row = ws.max_row
        max_col = ws.max_column

        ws['B2'] = series_names[i] + ' (10-YEAR ANALYSIS)'
        ws['B3'] = 'INLET TYPE'
        ws['C3'] = 'STRUCTURE'
        ws['E3'] = 'A'
        ws['F3'] = 'TC'
        ws['G3'] = 'I'
        ws['H3'] = 'C'
        ws['I3'] = 'Q (INLET)'
        ws['J3'] = 'Q (TOTAL)'
        ws['K3'] = 'Q (CAPACITY)'
        ws['L3'] = 'PIPE LENGTH'
        ws['M3'] = 'PIPE SIZE'
        ws['N3'] = 'MATERIAL'
        ws['O3'] = 'PIPE SLOPE'
        ws['P3'] = 'UPPER INV'
        ws['Q3'] = 'LOWER INV'
        ws['R3'] = 'RIM ELEV UP'
        ws['S3'] = 'RIM ELEV DOWN'

        if hgl_toggle:
            ws['T3'] = 'HGL UP'
            ws['U3'] = 'HGL DOWN'

        ws['B4'] = ''
        ws['C4'] = 'FROM'
        ws['D4'] = 'TO'
        ws['E4'] = '(AC)'
        ws['F4'] = '(MIN)'
        ws['G4'] = '(IN/HR)'
        ws['H4'] = ''
        ws['I4'] = '(CFS)'
        ws['J4'] = '(CFS)'
        ws['K4'] = '(CFS)'
        ws['L4'] = '(FT)'
        ws['M4'] = '(IN)'
        ws['N4'] = ''
        ws['O4'] = '(%)'
        ws['P4'] = '(FT)'
        ws['Q4'] = '(FT)'
        ws['R4'] = '(FT)'
        ws['S4'] = '(FT)'

        if hgl_toggle:
            ws['T4'] = '(FT)'
            ws['U4'] = '(FT)'

        # Replace line number with structure name
        structures = [cell.value for row in ws.iter_rows(
            min_row=5,
            max_row=max_row,
            min_col=3,
            max_col=3
        ) for cell in row]

        for row in ws.iter_rows(
            min_row=5,
            max_row=max_row,
            min_col=4,
            max_col=4
        ):
            for cell in row:
                if cell.value == 'OUT':
                    pass
                else:
                    line = int(cell.value)
                    cell.value = structures[line - 1]

        # Replace n value with material type
        for row in ws.iter_rows(
            min_row=5,
            max_row=max_row,
            min_col=14,
            max_col=14
        ):
            for cell in row:
                cell.value = 'RCP'

        # Merge cells
        if hgl_toggle:
            ws.merge_cells('B2:U2')
        else:
            ws.merge_cells('B2:S2')

        ws.merge_cells('C3:D3')

        # Set cell font
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
                cell.font = Font(size=10, name='Arial')

        for row in ws.iter_rows(min_row=2, max_row=4,
                                min_col=2, max_col=max_col):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='Arial')

        # Set cell background fill
        ws['B2'].fill = PatternFill(fgColor='BFBFBF', bgColor='BFBFBF',
                                    fill_type='solid')

        for row in ws.iter_rows(min_row=3, max_row=4,
                                min_col=2, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill(fgColor='D9D9D9', bgColor='D9D9D9',
                                        fill_type='solid')

        # Set cell border
        set_border(ws, borders.thin, 3, max_row, 2, max_col)
        set_border(ws, borders.med_top_left_bot, 2, 2, 2, 2)
        set_border(ws, borders.med_top_right_bot, 2, 2, max_col, max_col)
        set_border(ws, borders.med_tlcorner, 3, 3, 2, 2)
        set_border(ws, borders.med_left, 4, 4, 2, 2)
        set_border(ws, borders.med_trcorner, 3, 3, max_col, max_col)
        set_border(ws, borders.med_right, 4, 4, max_col, max_col)
        set_border(ws, borders.med_tlcorner, 5, 5, 2, 2)
        set_border(ws, borders.med_trcorner, 5, 5, max_col, max_col)
        set_border(ws, borders.med_top_bot, 2, 2, 3, max_col - 1)
        set_border(ws, borders.med_top, 5, 5, 3, max_col - 1)

        # Set cell border for single entry
        if max_row == 5:
            set_border(ws, borders.med_top_left_bot,
                       max_row, max_row, 2, 2)
            set_border(ws, borders.med_top_right_bot,
                       max_row, max_row, max_col, max_col)
            set_border(ws, borders.med_top_bot,
                       max_row, max_row, 3, max_col - 1)

        else:
            set_border(ws, borders.med_left,
                       6, max_row, 2, 2)
            set_border(ws, borders.med_right,
                       6, max_row, max_col, max_col)
            set_border(ws, borders.med_bot,
                       max_row, max_row, 3, max_col - 1)
            set_border(ws, borders.med_blcorner,
                       max_row, max_row, 2, 2)
            set_border(ws, borders.med_brcorner,
                       max_row, max_row, max_col, max_col)

        # Set cell number format
        set_format(ws, '0.00', 5, max_row, 5, 5)
        set_format(ws, '0.0', 5, max_row, 6, 6)
        set_format(ws, '0.00', 5, max_row, 7, 11)
        set_format(ws, '0', 5, max_row, 12, 13)
        set_format(ws, '0.00', 5, max_row, 15, max_col)

        # Set row dimensions
        row_dims = {1: 13.8, 2: 18, 3: 21, 4: 21}

        for key, value in row_dims.items():
            ws.row_dimensions[key].height = value

        # Set column dimensions
        col_dims = {'A': 4.02, 'B': 13.13, 'C': 10.02, 'D': 10.02, 'E': 9.58,
                    'F': 9.58, 'G': 9.58, 'H': 9.58, 'I': 12.24, 'J': 12.24,
                    'K': 15.24, 'L': 15.58, 'M': 11.47, 'N': 13.91, 'O': 14.24,
                    'P': 15.13, 'Q': 15.13, 'R': 17.24, 'S': 17.24, 'T': 13.8,
                    'U': 13.8}

        for key, value in col_dims.items():
            ws.column_dimensions[key].width = value

    # Convert formatted xlsx file to binary
    stream_xlsx_form = io.BytesIO()
    wb.save(stream_xlsx_form)
    stream_xlsx_form.seek(0)

    # Write formatted xlsx file to s3
    xlsx_form = stream_xlsx_form.getvalue()
    s3_key_xlsx = ''.join(['design', '/', folder_name, '/',
                           'xlsx', '/', 'Pipe Design.xlsx'])
    s3.put_object(Bucket=S3_BUCKET, Key=s3_key_xlsx, Body=xlsx_form)

    # Return s3 key
    response = s3_key_xlsx
    return response
